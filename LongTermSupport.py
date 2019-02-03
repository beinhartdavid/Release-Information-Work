#InplviewHashdc454d86-66cc-4c21-9b99-ceeeccd108dc=Paged%3DTRUE-p_GroupCol1%3D20150820%252004%253a00%253a00-PageFirstRow%3D131-WebPartID%3D%7B678EB118--B7F1--4806--8A26--8ACB90A7FE88%7Dhttps://collaboration.nhlbi.nih.gov/orgs/od/om/itac/sao/cab/rm/rpd/SitePages/Home.aspx
import datetime
import time
from selenium import webdriver
import pickle
import re
from openpyxl import load_workbook

url = ""
driver = webdriver.Chrome(executable_path='C:\\chromedriver.exe')  # Optional argument, if not specified will search path.
driver.get(url);
Results = []

#only care about the most recent panel
tabs = driver.find_elements_by_partial_link_text("Release Deployment Planning Date")
tabs[0].click()
time.sleep(2)  #Allow Each Page to Open
name = driver.find_elements_by_css_selector("[id^='tbod']")
for n in name:
    items = n.find_elements_by_tag_name("td")
    for i in items:
        print(i.text)
        Results.append(i.text)

print("DONE!!!!")
driver.quit()


with open("data/RPDLT.txt", "wb") as fp:
    pickle.dump(Results,fp)


with open("data/RPDLT.txt", "rb") as fp:
    data = pickle.load(fp)

lst = []
Final = []
for d in data:
    if d == "View Item":
        if lst == []: continue
        else:
            Final.append(lst)
            lst = []
            continue
    if d == "": continue
    lst.append(d)
    date = re.match("^[0-9]*[\/][0-9]*[\/][0-9]*$", d)
    if date:
        Final.append(lst)
        lst = []
for f in Final:
    if len(f) > 10: print(f)

# Write to  Excel:
Columns = ["A", "B", "C", "D", "E", "F", "G", "H","I","J"]
Results = ["1", "2", "3", "4", "5", "6", "7", "8", "9", "10"]

wb = load_workbook("data/RPDLT.xlsx")

H = wb["Sheet1"]

index = H.max_row + 1
# Make Header


today = datetime.date.today()
H[Columns[0] + str(index)] = "Testing on " + str(today)
index += 1

for r in Final:
    # if len(r.keys()) != 6: continue
    for i in range(len(r)):
        H[Columns[i] + str(index)] = r[i]
    index += 1

today = datetime.date.today()
H[Columns[0] + str(index)] = "Check for " + str(today) + " Complete"
wb.save("data/RPDLT.xlsx")